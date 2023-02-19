"""
5. Прочитать сохранённый csv-файл и сохранить данные в excel-файл,
кроме возраста - столбец с этими данными не нужен.
"""

import openpyxl
import csv

with open('task_4.csv') as csv_file:
    wb = openpyxl.Workbook()
    wb.create_sheet('Page 1', 0)
    ws = wb['Page 1']
    for index, row in enumerate(csv.reader(csv_file, delimiter=',')):
        row.pop(2)
        if index != 0:  # person must be at least one
            ws.cell(row=1, column=index + 1, value=f'Person {index}')
        for value_index, value in enumerate(row):
            ws.cell(row=value_index + 2, column=index + 1, value=value)

    wb.save('task_5.xlsx')
