from datetime import datetime
from sys import stderr
from time import sleep
from typing import Optional, Any, Iterable, TYPE_CHECKING
if TYPE_CHECKING:
    from data.data_access import GeneratorTypes

import openpyxl

from data import (
    db_provider,
    CATEGORY_DB_NAME,
    CATEGORY_DB_TYPE,
    PRODUCTS_DB_NAME,
    PRODUCTS_DB_TYPE,
    ORDERS_DB_NAME,
    ORDERS_DB_TYPE,
    ProductExistsError,
)
from validators import (
    validate_date,
    validate_category_name,
    validate_parameter_name,
    validate_product_name,
    validate_quantity,
    validate_price,
    validate_parameter_value,
    validate_category_exists,
    validate_id
)

category_db = db_provider(data_name=CATEGORY_DB_NAME,
                          data_type=CATEGORY_DB_TYPE)
products_db = db_provider(data_name=PRODUCTS_DB_NAME,
                          data_type=PRODUCTS_DB_TYPE)
orders_db = db_provider(data_name=ORDERS_DB_NAME,
                        data_type=ORDERS_DB_TYPE)


class FilterMixin:
    @staticmethod
    def category_filter(param: str, category: str) -> bool:
        return not param or category == param

    @staticmethod
    def min_date_filter(param: str, date: str) -> bool:
        if param == '':
            return True
        param_obj = datetime.strptime(param, '%Y-%m-%d')
        date_obj = datetime.strptime(date[:10], '%Y-%m-%d')
        return not param or param_obj <= date_obj

    @staticmethod
    def max_date_filter(param: str, date: str) -> bool:
        if param == '':
            return True
        param_obj = datetime.strptime(param, '%Y-%m-%d')
        date_obj = datetime.strptime(date[:10], '%Y-%m-%d')
        return not param or param_obj >= date_obj


class Category:
    def __init__(self) -> None:
        self.parameters: list[str] = list()
        self.category_db = category_db

    def __call__(self, *args: Any, **kwargs: Any) -> None:
        self.get_parameters()
        self.add_new_category()

    @property
    def category_name(self) -> str:
        return self._category_name

    @category_name.setter
    def category_name(self, value: str) -> None:
        if value != '':
            validate_category_name(category_name=value)
        self._category_name = value

    def get_parameters(self) -> None:
        counter = 0
        while counter <= 10:
            parameter = input('Enter the parameter name: ')
            validate_parameter_name(parameter_name=parameter)
            if parameter == 'stop':
                break
            self.parameters.append(parameter.strip())
            counter += 1

    def add_new_category(self) -> None:
        record = {
            self.category_name: self.parameters
        }
        self.category_db.create(record=record)

    def categories_read(self) -> str:
        categories = [category for category in self.category_db.read()]
        return '\n'.join(categories)

    def parameters_read(self, category_name: str) -> list[str]:
        parameters: list[str] = self.category_db.read_all()[category_name]
        return parameters


class Product(Category, FilterMixin):
    def __init__(self) -> None:
        super().__init__()
        self.products_db = products_db
        self.orders_db = orders_db
        self.parameters_dict: dict[str, Any] = dict()

    def __call__(self, *args: Any, **kwargs: Any) -> None:
        self.get_parameters_value()
        self.add_product()

    @property
    def product_name(self) -> str:
        return self._product_name

    @product_name.setter
    def product_name(self, value: str) -> None:
        validate_product_name(product_name=value)
        self._product_name = value

    @property
    def quantity(self) -> int:
        return self._quantity

    @quantity.setter
    def quantity(self, value: str) -> None:
        validate_quantity(quantity=value)
        self._quantity = int(value)

    @property
    def price(self) -> float:
        return self._price

    @price.setter
    def price(self, value: str) -> None:
        validate_price(price=value)
        self._price = float(value)

    def get_parameters_value(self) -> None:
        validate_category_exists(category_name=self.category_name,
                                 categories=self.categories_read())
        for parameter in self.parameters_read(
                category_name=self.category_name):
            parameter_value = input(f'Enter a value for the '
                                    f'{parameter} parameter: ')
            validate_parameter_value(parameter=parameter_value)
            self.parameters_dict[parameter] = parameter_value

    def add_product(self) -> None:
        record: dict[str, Any] = {
            'name': self.product_name,
            'category': self.category_name,
        }
        record.update(self.parameters_dict)
        record.update(
            {
                'quantity': self.quantity,
                'price': self.price,
                'created_at': str(
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                'updated_at': str(
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            }
        )
        id_exists = self.products_db.id_exists_to_update(record=record)
        if id_exists is None:
            self.products_db.create(record=record)
        else:
            product_id = id_exists
            self.products_db.update(product_id=product_id, record=record)

    def delete_product(self, product_id: str) -> None:
        validate_id(product_id=product_id)
        self.products_db.validate_product_exists(product_id=product_id)
        self.products_db.delete(product_id=product_id)

    def get_products_list(self, min_date: str, max_date: str) -> str:
        validate_date(min_date)
        validate_date(max_date)
        result = ''
        product: tuple[str, dict[str, Any]]
        for product in self.products_db.read():
            if self.category_filter(
                    param=self.category_name,
                    category=product[1]['category']) and \
                    self.min_date_filter(
                        param=min_date,
                        date=product[1]['created_at']) and \
                    self.max_date_filter(
                        param=max_date,
                        date=product[1]['created_at']):
                result += f'id: {product[0]} | name: {product[1]["name"]}\n'
        return result

    def get_product(self, product_id: str) -> str:
        validate_id(product_id=product_id)
        try:
            self.products_db.validate_product_exists(product_id=product_id)
            products = self.products_db.read_all()['Goods']
            product = products.get(product_id)
            result = f'ID {product_id} | '
            for key, value in product.items():
                result += f'{key}: {value} | '
            result += 'You can buy it now!'
            return result
        except ProductExistsError as err:
            print(err, file=stderr)
            sleep(0.5)
            return "You can't buy it."

    def products_order(self, products_id: dict[str, Any]) -> str:
        order = 'Your order:\n'
        total_price = 0.0
        products = self.products_db.read_all()['Goods']
        data_update = dict()
        order_dict: dict[str, Any] = {
            'Product list': dict(),
            'Total price': 0.0,
            'Created_at': str(datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        }
        for product_id, quantity in products_id.items():
            try:
                self.products_db.validate_product_exists(product_id=product_id)
            except ProductExistsError as err:
                print(err, file=stderr)
                sleep(0.5)
                return 'Order canceled.'

            if products[product_id]['quantity'] >= quantity:
                order += f'ID: {product_id} | ' \
                         f'Product name: {products[product_id]["name"]} ' \
                         f'| Price: {products[product_id]["price"]} ' \
                         f'| Quantity: {quantity}\n'
                total_price += products[product_id]["price"] * quantity
                data_update[product_id] = {
                    'quantity': products[product_id]['quantity'] - quantity,
                    'updated_at':
                        str(datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                }
                order_dict['Product list'][product_id] = (
                    products[product_id]['name'],
                    products[product_id]['category'],
                    quantity,
                    products[product_id]['price'],
                )

            else:
                order = f'Not enough product with ID {product_id} in ' \
                        f'stock (quantity ' \
                        f'{products[product_id]["quantity"]}). ' \
                        f'Order canceled.\n'
                return order

        order += f'Total price: {total_price}'
        order_dict['Total price'] = total_price
        for product_id, record in data_update.items():
            if record['quantity'] == 0:
                self.products_db.delete(product_id=product_id)
            else:
                self.products_db.update(product_id=product_id, record=record)
        self.orders_db.create(record=order_dict)
        return order


class CreateStatistics(FilterMixin):
    def __init__(self, min_date: str, max_date: str) -> None:
        self.min_date = min_date
        self.max_date = max_date
        self.workbook = openpyxl.Workbook()
        self.sheet_categories = self.workbook.active
        self.sheet_products = self.workbook.create_sheet('Products')
        self.sheet_orders = self.workbook.create_sheet('Orders')
        self.sheet_metrics = self.workbook.create_sheet('Metrics')
        self.data_orders = orders_db.read_all()
        self.categories = {category: [0, 0] for category in category_db.read()}
        self.products = {params["name"]: 0 for product_id, params in
                         products_db.read()}
        self.orders: dict[str, Any] = dict()
        self.metrics = {
            'Total revenue': 0,
            'Total quantity': 0,
            'Popular category': '-',
            'Popular product': '-',
        }

    def __call__(self, *args: Any, **kwargs: Any) -> None:
        self.update_statistics()
        self.create_sheet_category()
        self.create_sheet_products()
        self.create_sheet_orders()
        self.create_sheet_metrics()
        self.save_excel_file()

    @property
    def min_date(self) -> str:
        return self._min_date

    @min_date.setter
    def min_date(self, date: str) -> None:
        validate_date(date)
        self._min_date = date

    @property
    def max_date(self) -> str:
        return self._max_date

    @max_date.setter
    def max_date(self, date: str) -> None:
        validate_date(date)
        self._max_date = date

    def update_statistics(self) -> None:
        for order_id, order in self.data_orders.items():
            if self.min_date_filter(param=self.min_date,
                                    date=order['Created_at']) and \
                    self.max_date_filter(param=self.max_date,
                                         date=order['Created_at']):
                self.orders[order_id] = order['Total price']
                self.metrics['Total revenue'] += order['Total price']
                for product_id, product in order['Product list'].items():
                    if product[0] not in self.products:
                        self.products[product[0]] = 0
                    self.categories[product[1]][0] += product[2]
                    self.categories[product[1]][1] += product[3] * product[2]
                    self.products[product[0]] += product[2]
                    self.metrics['Total quantity'] += product[2]

    def create_sheet_category(self) -> None:
        self.sheet_categories.title = 'Categories'
        self.sheet_categories.cell(row=1, column=1).value = 'Categories'
        self.sheet_categories.cell(row=1, column=2).value = 'Quantity'
        self.sheet_categories.cell(row=1, column=3).value = 'Total revenue'
        self.sheet_categories.column_dimensions['A'].width = 15
        self.sheet_categories.column_dimensions['C'].width = 15

        for index, category in enumerate(self.categories):
            self.sheet_categories.cell(row=index + 2,
                                       column=1).value = category
            self.sheet_categories.cell(row=index + 2, column=2).value = \
                self.categories[category][0]
            self.sheet_categories.cell(row=index + 2, column=3).value = \
                self.categories[category][1]

    def create_sheet_products(self) -> None:
        self.sheet_products.cell(row=1, column=1).value = 'Name'
        self.sheet_products.cell(row=1, column=2).value = 'Quantity'
        self.sheet_products.column_dimensions['A'].width = 30

        for index, product in enumerate(self.products):
            self.sheet_products.cell(row=index + 2, column=1).value = product
            self.sheet_products.cell(row=index + 2, column=2).value = \
                self.products[
                    product]

    def create_sheet_orders(self) -> None:
        self.sheet_orders.cell(row=1, column=1).value = 'ID'
        self.sheet_orders.cell(row=1, column=2).value = 'Total revenue'
        self.sheet_orders.column_dimensions['B'].width = 15

        sorted_orders = dict(
            sorted(self.orders.items(), key=lambda item: item[1],
                   reverse=True))

        for index, order_id in enumerate(sorted_orders):
            self.sheet_orders.cell(row=index + 2, column=1).value = order_id
            self.sheet_orders.cell(row=index + 2, column=2).value = \
                sorted_orders[order_id]

    def create_sheet_metrics(self) -> None:
        self.sheet_metrics.cell(row=1, column=1).value = 'Total revenue'
        self.sheet_metrics.cell(row=1, column=2).value = 'Total quantity'
        self.sheet_metrics.cell(row=1,
                                column=3).value = 'Most popular category'
        self.sheet_metrics.cell(row=1, column=4).value = 'Most popular product'
        self.sheet_metrics.column_dimensions['A'].width = 15
        self.sheet_metrics.column_dimensions['B'].width = 15
        self.sheet_metrics.column_dimensions['C'].width = 20
        self.sheet_metrics.column_dimensions['D'].width = 30
        sorted_products = dict(
            sorted(self.products.items(), key=lambda item: item[1],
                   reverse=True))
        sorted_categories = dict(
            sorted(self.categories.items(), key=lambda item: item[1],
                   reverse=True))
        if len(self.orders) != 0:
            self.metrics['Popular category'] = list(sorted_categories.keys())[
                0]
            self.metrics['Popular product'] = list(sorted_products.keys())[0]

        for index, metric_id in enumerate(self.metrics):
            self.sheet_metrics.cell(row=2, column=index + 1).value = \
                self.metrics[
                    metric_id]

    def save_excel_file(self) -> None:
        if not self.min_date and not self.max_date:
            self.workbook.save('statistics.xlsx')
        else:
            self.workbook.save(
                f'statistics_{self.min_date}-{self.max_date}.xlsx')
